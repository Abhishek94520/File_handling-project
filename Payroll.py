import os
from datetime import datetime
import openpyxl
import shutil


def copyFile(src, dst):
	shutil.copy(src, dst)

def deleteFile(file):
	os.remove(file)

def getDate():
	format = '%Y-%m-%d'
	date = datetime.today().strftime(format)
	return date

def findBaseSalary(toProcess):  #this fuction will return the dictionary with empid as key and value = [working_days, base_Salary]
	empDetails = 'demo.xlsx'

	wb = openpyxl.load_workbook(empDetails)
	sheet = wb.active

	row = sheet.max_row
	column = sheet.max_column

	for i in range(2, row + 1):
		cur = sheet.cell(row = i, column = 1).value
		
		if cur in toProcess:
			base_salary =  sheet.cell(row = i, column = column).value
			toProcess[cur].append(base_salary)
	
	return toProcess

def createPending(pending):
	if pending == []:
		file = "pending/pending.xlsx"
		headers = ['Employee ID']
		wb = openpyxl.Workbook()
		sheet = wb.active
		sheet.append(headers)
		for empId in pending:
			sheet.append([empId])
		wb.save(file)

def processPayroll(toProcess):
	date = getDate()
	file = 'Employee_salary_' + date + "_.xlsx"
	path = 'payroll/' + file

	pending = []

	print('Creating file ', file)
	headers = ['Employee ID', 'Working Days', 'Salary']
	wb = openpyxl.Workbook()
	sheet = wb.active
	sheet.append(headers)

	for empId, value in toProcess.items():
		if len(value) < 2:
			pending.append(empId)
		else:
			working_days = value[0]
			base_salary = value[1]

			cur_salary = (base_salary / 30) * working_days

			row = [empId, working_days, cur_salary]
			sheet.append(row)

	wb.save(path)

	dst = 'share/' + file

	print('Moving file from', path , ' to ', dst)
	copyFile(path, dst)
	deleteFile(path)
	

	print('Creating Pending.xlsx file. ')
	print(f'.>>>>>>>>{pending}<<<<<')

	createPending(pending)
	

#MAIN 
def main():
	now = datetime.now()
	cur_date = int(now.strftime("%d"))
	cur_date = 21

	if cur_date > 20:  #only process when current date > 20
		date = getDate()
		src = 'account/Payroll.txt'
		dst = 'work/' + date + '.txt'

		print('Copying File from ', src , ' to ', dst)
		copyFile(src, dst)



		file = dst    # The file in the work folder

		toProcess = {}  # Dictionary of employee's Id whose payroll is to process

		print('Reading file ', dst)
		fs = open(file, 'r')
		for line in fs:
			line = line.replace('empcode:', "")
			line = line.replace('work_days:', "")

			empId, working_days = line.split()
			working_days = int(working_days)

			toProcess[empId] = [working_days]

		fs.close()


		if toProcess: 

			print('Collecting neccessary details.')
			toProcess = findBaseSalary(toProcess)
			print('Processing Payroll')
			processPayroll(toProcess)

		else:
			print('No Payroll to be processed.')

if __name__=="__main__":
	main()
